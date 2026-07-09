import json
import unittest
from datetime import datetime
from pathlib import Path
from tempfile import TemporaryDirectory
from types import SimpleNamespace
from unittest.mock import Mock, patch

from openpyxl import load_workbook

from local_actions import action_log
from local_actions import actions as main
from local_actions import cli, direct_commands, registry, slm, uia_agent


class UrlActionTests(unittest.TestCase):
    def test_google_search_encodes_japanese_query(self) -> None:
        self.assertEqual(
            main.google_search("生成 AI"),
            "https://www.google.com/search?q=%E7%94%9F%E6%88%90+AI",
        )

    def test_x_search_encodes_japanese_query(self) -> None:
        self.assertEqual(
            main.x_open("search", "生成 AI"),
            "https://x.com/search?q=%E7%94%9F%E6%88%90+AI&src=typed_query",
        )

    def test_x_open_supports_fixed_destinations(self) -> None:
        self.assertEqual(main.x_open("home"), "https://x.com/home")
        with patch.dict(
            main.os.environ,
            {main.X_USERNAME_ENV_VAR: "@OpenAI"},
            clear=True,
        ):
            self.assertEqual(
                main.x_open("likes"),
                "https://x.com/OpenAI/likes",
            )
            self.assertEqual(
                main.x_open("profile", "OtherUser"),
                "https://x.com/OpenAI",
            )

    def test_x_open_requires_configured_username(self) -> None:
        with (
            patch.dict(main.os.environ, {}, clear=True),
            self.assertRaisesRegex(OSError, main.X_USERNAME_ENV_VAR),
        ):
            main.x_open("profile")

    def test_x_open_rejects_invalid_configured_username(self) -> None:
        with (
            patch.dict(
                main.os.environ,
                {main.X_USERNAME_ENV_VAR: "invalid/name"},
                clear=True,
            ),
            self.assertRaisesRegex(ValueError, "ユーザー名が不正"),
        ):
            main.x_open("profile")


class CaptureTests(unittest.TestCase):
    def test_capture_creates_headers_and_appends_rows(self) -> None:
        with TemporaryDirectory() as directory:
            capture_path = Path(directory) / "inbox.xlsx"
            with patch(
                "local_actions.actions.get_capture_path",
                return_value=capture_path,
            ):
                result = main.capture("memo", " 日本語のメモ ", "本文")
                main.capture("log", "作業記録")

            workbook = load_workbook(capture_path, read_only=True)
            rows = list(workbook.active.iter_rows(values_only=True))
            workbook.close()

        self.assertEqual(rows[0], main.CAPTURE_HEADERS)
        self.assertEqual(rows[1][1:], ("memo", "日本語のメモ", "本文"))
        self.assertEqual(rows[2][1:], ("log", "作業記録", None))
        self.assertIn("Excel INBOXへ保存しました", result)

    def test_capture_rejects_empty_title_and_unknown_kind(self) -> None:
        with self.assertRaisesRegex(ValueError, "タイトルが空"):
            main.capture("memo", " ")
        with self.assertRaisesRegex(ValueError, "未対応のキャプチャ種別"):
            main.capture("other", "題名")


class CurrentPageTests(unittest.TestCase):

    @patch("local_actions.actions.subprocess.run")
    def test_get_current_page_parses_ui_automation_result(
        self,
        run: Mock,
    ) -> None:
        run.return_value = Mock(
            returncode=0,
            stdout='{"title":"Example","url":"https://example.com"}',
            stderr="",
        )

        with patch("local_actions.actions.sys.platform", "win32"):
            self.assertEqual(
                main.get_current_page(),
                "Example\nhttps://example.com",
            )

        command = run.call_args.args[0]
        self.assertEqual(command[:4], [
            "powershell.exe",
            "-NoProfile",
            "-NonInteractive",
            "-Command",
        ])

    @patch("local_actions.actions.subprocess.run")
    def test_get_current_page_reports_ui_automation_failure(
        self,
        run: Mock,
    ) -> None:
        run.return_value = Mock(
            returncode=1,
            stdout="",
            stderr="Browser address bar was not found.",
        )

        with (
            patch("local_actions.actions.sys.platform", "win32"),
            self.assertRaisesRegex(OSError, "ページ情報を取得できません"),
        ):
            main.get_current_page()


class OneDriveTests(unittest.TestCase):
    def test_get_onedrive_directory_uses_consumer_folder(self) -> None:
        with TemporaryDirectory() as directory:
            with (
                patch("local_actions.actions.sys.platform", "win32"),
                patch.dict(
                    main.os.environ,
                    {"OneDriveConsumer": directory},
                    clear=True,
                ),
            ):
                self.assertEqual(
                    main.get_onedrive_directory(),
                    Path(directory) / "Local Actions",
                )

    def test_get_onedrive_directory_reports_missing_configuration(self) -> None:
        with (
            patch("local_actions.actions.sys.platform", "win32"),
            patch.dict(main.os.environ, {}, clear=True),
            self.assertRaisesRegex(OSError, "OneDriveフォルダを特定"),
        ):
            main.get_onedrive_directory()

class CalendarTaskTests(unittest.TestCase):
    def test_format_calendar_ics_includes_escaped_body(self) -> None:
        ics = main.format_calendar_ics(
            "【タスク】田中さんに返信",
            "2026-07-01T14:00:00",
            "メモ; 重要, 明日",
            "uid@dev-slm-shortcut",
            "20260701T050000Z",
        )

        self.assertIn("SUMMARY:【タスク】田中さんに返信", ics)
        self.assertIn("DTSTART:20260701T140000", ics)
        self.assertIn("DTEND:20260701T141500", ics)
        self.assertIn(r"DESCRIPTION:メモ\; 重要\, 明日", ics)

    def test_format_calendar_ics_omits_empty_description(self) -> None:
        ics = main.format_calendar_ics(
            "【リマインダ】水を飲む",
            "2026-07-01T14:00:00",
            "",
            "uid@dev-slm-shortcut",
            "20260701T050000Z",
        )

        event = ics.split("BEGIN:VALARM")[0]
        self.assertNotIn("DESCRIPTION:", event)

    @patch("local_actions.actions.time.sleep")
    @patch("local_actions.actions.os.startfile")
    def test_create_calendar_task_opens_and_cleans_up(
        self,
        startfile: Mock,
        sleep: Mock,
    ) -> None:
        with patch("local_actions.actions.sys.platform", "win32"):
            result = main.create_calendar_task(
                "2026-07-01T14:00:00",
                "タスク",
                "田中さんに返信",
            )

        opened_path = Path(startfile.call_args.args[0])
        self.assertEqual(opened_path.suffix, ".ics")
        self.assertFalse(opened_path.exists())
        self.assertIn("【タスク】田中さんに返信", result)

    def test_create_calendar_task_rejects_empty_title(self) -> None:
        with (
            patch("local_actions.actions.sys.platform", "win32"),
            self.assertRaisesRegex(ValueError, "タスク名が空"),
        ):
            main.create_calendar_task("2026-07-01T14:00:00", "タスク", "  ")

    def test_create_calendar_task_rejects_unknown_entity_type(self) -> None:
        with (
            patch("local_actions.actions.sys.platform", "win32"),
            self.assertRaisesRegex(ValueError, "未対応の種別"),
        ):
            main.create_calendar_task(
                "2026-07-01T14:00:00",
                "予定",
                "田中さんに返信",
            )


class WindowsActionTests(unittest.TestCase):
    @patch("local_actions.actions.os.startfile")
    def test_open_folder_opens_downloads_folder(
        self,
        startfile: Mock,
    ) -> None:
        with TemporaryDirectory() as directory:
            home = Path(directory)
            downloads = home / "Downloads"
            downloads.mkdir()
            with (
                patch("local_actions.actions.sys.platform", "win32"),
                patch("local_actions.actions.Path.home", return_value=home),
            ):
                result = main.open_folder("downloads")

        startfile.assert_called_once_with(downloads)
        self.assertIn(str(downloads), result)

    @patch("local_actions.actions.os.startfile")
    def test_open_folder_opens_onedrive_save_directory(
        self,
        startfile: Mock,
    ) -> None:
        with TemporaryDirectory() as directory:
            onedrive = Path(directory)
            save_directory = onedrive / "Local Actions"
            save_directory.mkdir()
            with (
                patch("local_actions.actions.sys.platform", "win32"),
                patch.dict(
                    main.os.environ,
                    {"OneDriveConsumer": str(onedrive)},
                    clear=True,
                ),
            ):
                result = main.open_folder("onedrive")

        startfile.assert_called_once_with(save_directory)
        self.assertIn(str(save_directory), result)

    def test_open_folder_rejects_unknown_folder(self) -> None:
        with (
            patch("local_actions.actions.sys.platform", "win32"),
            self.assertRaisesRegex(ValueError, "未対応のフォルダ"),
        ):
            main.open_folder("arbitrary")  # type: ignore[arg-type]

    @patch("local_actions.actions.os.startfile")
    def test_open_settings_uses_allowlisted_uri(self, startfile: Mock) -> None:
        with patch("local_actions.actions.sys.platform", "win32"):
            result = main.open_settings("windows_update")

        startfile.assert_called_once_with("ms-settings:windowsupdate")
        self.assertIn("Windows Update", result)

    def test_open_settings_rejects_unknown_page(self) -> None:
        with (
            patch("local_actions.actions.sys.platform", "win32"),
            self.assertRaisesRegex(ValueError, "未対応の設定ページ"),
        ):
            main.open_settings("arbitrary")  # type: ignore[arg-type]

    @patch("local_actions.actions.run_clipboard_script")
    def test_copy_text_passes_text_without_modification(
        self,
        run_script: Mock,
    ) -> None:
        text = "日本語\nをそのまま"
        result = main.copy_text(text)

        run_script.assert_called_once_with(main.COPY_TEXT_SCRIPT, text)
        self.assertIn(f"{len(text)}文字", result)

    @patch(
        "local_actions.actions.run_clipboard_script",
        return_value="コピー済みの内容",
    )
    def test_get_clipboard_text_returns_script_output(
        self,
        run_script: Mock,
    ) -> None:
        self.assertEqual(main.get_clipboard_text(), "コピー済みの内容")
        run_script.assert_called_once_with(main.GET_CLIPBOARD_TEXT_SCRIPT)

    @patch("local_actions.actions.shutil.disk_usage")
    def test_show_system_info_formats_windows_and_disk_data(
        self,
        disk_usage: Mock,
    ) -> None:
        disk_usage.return_value = Mock(
            total=100 * 1024 ** 3,
            used=40 * 1024 ** 3,
            free=60 * 1024 ** 3,
        )
        windows_version = Mock(build=26100)

        with (
            patch("local_actions.actions.sys.platform", "win32"),
            patch(
                "local_actions.actions.sys.getwindowsversion",
                return_value=windows_version,
            ),
            patch(
                "local_actions.actions.platform.win32_ver",
                return_value=("11", "10.0", "", ""),
            ),
            patch(
                "local_actions.actions.platform.win32_edition",
                return_value="Professional",
            ),
            patch(
                "local_actions.actions.platform.node",
                return_value="TEST-PC",
            ),
            patch.dict(main.os.environ, {"SystemDrive": "D:"}),
        ):
            result = main.show_system_info()

        disk_usage.assert_called_once_with("D:\\")
        self.assertIn("Windows Professional 11", result)
        self.assertIn("ビルド 26100", result)
        self.assertIn("空き容量: 60.0 GiB", result)

    @patch(
        "local_actions.actions.ctypes.windll.user32.LockWorkStation",
        return_value=1,
    )
    def test_lock_pc_runs_without_confirmation(self, lock: Mock) -> None:
        with patch("local_actions.actions.sys.platform", "win32"):
            self.assertEqual(main.lock_pc(), "PCをロックしました。")

        lock.assert_called_once_with()
        self.assertIsNone(registry.actions["lock_pc"].confirmation_message)


class SelectActionTests(unittest.TestCase):
    @patch("local_actions.slm.chat")
    def test_select_actions_preserves_structured_step_order(
        self,
        chat: Mock,
    ) -> None:
        chat.return_value = Mock(
            message=Mock(
                thinking="登録済みActionと引数を検討しました。",
                content=json.dumps(
                    {
                        "steps": [
                            {
                                "action": "get_clipboard_text",
                                "arguments": {},
                            },
                            {
                                "action": "create_calendar_task",
                                "arguments": {
                                    "start_time": "2026-07-03T15:00:00",
                                    "entity_type": "タスク",
                                    "title": "田中さんに返信",
                                },
                            },
                        ]
                    },
                    ensure_ascii=False,
                )
            ),
        )

        with patch("builtins.print") as output:
            self.assertEqual(
                slm.select_actions("クリップボードの内容と一緒に登録"),
                [
                    registry.PlannedAction("get_clipboard_text", {}),
                    registry.PlannedAction(
                        "create_calendar_task",
                        {
                            "start_time": "2026-07-03T15:00:00",
                            "entity_type": "タスク",
                            "title": "田中さんに返信",
                        },
                    ),
                ],
            )

        self.assertFalse(chat.call_args.kwargs["think"])
        output.assert_called_once_with("登録済みActionと引数を検討しました。")

    def test_plan_schema_exposes_body_as_optional_argument(self) -> None:
        schema = slm.build_action_plan_schema()
        step_schemas = schema["properties"]["steps"]["items"]["oneOf"]
        calendar_schema = next(
            step
            for step in step_schemas
            if step["properties"]["action"]["const"]
            == "create_calendar_task"
        )
        arguments_schema = calendar_schema["properties"]["arguments"]
        argument_properties = arguments_schema["properties"]

        self.assertIn("title", argument_properties)
        self.assertIn("body", argument_properties)
        self.assertNotIn("body", arguments_schema.get("required", []))

    def test_plan_schema_exposes_capture_kind_and_optional_body(self) -> None:
        schema = slm.build_action_plan_schema()
        step_schemas = schema["properties"]["steps"]["items"]["oneOf"]
        capture_schema = next(
            step
            for step in step_schemas
            if step["properties"]["action"]["const"] == "capture"
        )
        arguments_schema = capture_schema["properties"]["arguments"]

        self.assertEqual(
            arguments_schema["properties"]["kind"]["enum"],
            ["memo", "log"],
        )
        self.assertIn("title", arguments_schema["required"])
        self.assertNotIn("body", arguments_schema["required"])


class CalendarScheduleTests(unittest.TestCase):
    def test_default_start_uses_lunch_slot_in_morning(self) -> None:
        self.assertEqual(
            slm.default_calendar_start(datetime(2026, 7, 3, 9, 0)),
            datetime(2026, 7, 3, 12, 15),
        )

    def test_default_start_uses_afternoon_slot_before_evening(self) -> None:
        self.assertEqual(
            slm.default_calendar_start(datetime(2026, 7, 3, 13, 30)),
            datetime(2026, 7, 3, 15, 0),
        )

    def test_default_start_moves_to_next_business_day_at_night(self) -> None:
        # 2026-07-03は金曜日のため、翌営業日は月曜の2026-07-06。
        self.assertEqual(
            slm.default_calendar_start(datetime(2026, 7, 3, 20, 0)),
            datetime(2026, 7, 6, 8, 45),
        )


class WorkflowTests(unittest.TestCase):
    def test_workflow_injects_previous_result_into_registered_argument(
        self,
    ) -> None:
        calls = Mock()

        def clipboard() -> str:
            calls("clipboard")
            return "https://teams.example/message/123"

        def calendar(title: str, body: str = "") -> str:
            calls("calendar", title=title, body=body)
            return "[ダミー] 登録しました。"

        plan = [
            registry.PlannedAction("get_clipboard_text", {}),
            registry.PlannedAction(
                "create_calendar_task",
                {"title": "田中さんに返信"},
            ),
        ]

        with (
            patch.dict(
                registry.actions,
                {
                    "get_clipboard_text": registry.Action(clipboard),
                    "create_calendar_task": registry.Action(
                        calendar,
                        accepts_previous_as="body",
                    ),
                },
            ),
            patch("local_actions.registry.print") as print_function,
        ):
            result = registry.execute_workflow(plan)

        self.assertEqual(calls.call_args_list[0].args, ("clipboard",))
        self.assertEqual(calls.call_args_list[1].args, ("calendar",))
        self.assertEqual(
            calls.call_args_list[1].kwargs,
            {
                "title": "田中さんに返信",
                "body": "https://teams.example/message/123",
            },
        )
        print_function.assert_called_once_with("[ダミー] 登録しました。")
        self.assertEqual(result.result, "[ダミー] 登録しました。")

    def test_workflow_validates_all_steps_before_execution(self) -> None:
        clipboard = Mock(return_value="機密情報")
        plan = [
            registry.PlannedAction("get_clipboard_text", {}),
            registry.PlannedAction("lock_pc", {}),
        ]

        with (
            patch.dict(
                registry.actions,
                {"get_clipboard_text": registry.Action(clipboard)},
            ),
            self.assertRaisesRegex(ValueError, "受け取れません"),
        ):
            registry.execute_workflow(plan)

        clipboard.assert_not_called()

    def test_workflow_merges_model_body_with_previous_result(self) -> None:
        calls = Mock()

        def clipboard() -> str:
            return "https://teams.example/message/123"

        def calendar(title: str, body: str = "") -> str:
            calls(title=title, body=body)
            return "登録しました。"

        plan = [
            registry.PlannedAction("get_clipboard_text", {}),
            registry.PlannedAction(
                "create_calendar_task",
                {"title": "返信", "body": "モデルが書いた本文"},
            ),
        ]

        with (
            patch.dict(
                registry.actions,
                {
                    "get_clipboard_text": registry.Action(clipboard),
                    "create_calendar_task": registry.Action(
                        calendar,
                        accepts_previous_as="body",
                    ),
                },
            ),
            patch("local_actions.registry.print"),
        ):
            registry.execute_workflow(plan)

        self.assertEqual(
            calls.call_args.kwargs,
            {
                "title": "返信",
                "body": "モデルが書いた本文\n===\nhttps://teams.example/message/123",
            },
        )

    def test_workflow_combines_passthrough_results_before_injection(
        self,
    ) -> None:
        capture_call = Mock()

        def clipboard() -> str:
            return "クリップボード"

        def current_page() -> str:
            return "ページタイトル\nhttps://example.com"

        def capture(title: str, body: str = "") -> str:
            capture_call(title=title, body=body)
            return "保存しました。"

        plan = [
            registry.PlannedAction("get_clipboard_text", {}),
            registry.PlannedAction("get_current_page", {}),
            registry.PlannedAction("capture", {"title": "資料"}),
        ]

        with (
            patch.dict(
                registry.actions,
                {
                    "get_clipboard_text": registry.Action(
                        clipboard,
                        passthrough=True,
                    ),
                    "get_current_page": registry.Action(
                        current_page,
                        passthrough=True,
                    ),
                    "capture": registry.Action(
                        capture,
                        accepts_previous_as="body",
                    ),
                },
            ),
            patch("local_actions.registry.print"),
        ):
            registry.execute_workflow(plan)

        capture_call.assert_called_once_with(
            title="資料",
            body=(
                "クリップボード\n===\n"
                "ページタイトル\nhttps://example.com"
            ),
        )


class CommandLineOptionTests(unittest.TestCase):
    def test_format_action_list_contains_registered_actions(self) -> None:
        result = registry.format_action_list()

        self.assertIn(f"利用できる操作（{len(registry.actions)}件）", result)
        self.assertIn("google_maps_search(query)", result)
        self.assertIn("x_open(destination, query)", result)
        self.assertIn("open_folder(folder)", result)
        self.assertIn("get_current_page()", result)
        self.assertIn("capture(kind, title, body)", result)
        self.assertIn("copilot_chat(message, model)", result)
        self.assertNotIn("save_current_page()", result)
        self.assertNotIn("google_search(query)", result)
        self.assertNotIn("empty_recycle_bin()", result)

    @patch("local_actions.cli.select_actions")
    @patch("local_actions.cli.print")
    def test_list_option_does_not_call_ollama(
        self,
        print_function: Mock,
        select_actions: Mock,
    ) -> None:
        with patch("local_actions.cli.sys.argv", ["main.py", "--list"]):
            cli.main()

        select_actions.assert_not_called()
        listing = print_function.call_args.args[0]
        self.assertIn(registry.format_action_list(), listing)
        self.assertIn("tmp削除", listing)
        self.assertIn("WSL圧縮", listing)


class DirectCommandTests(unittest.TestCase):
    def test_direct_command_aliases_use_exact_matching(self) -> None:
        self.assertEqual(
            direct_commands.match_direct_command("tmp削除").operation,
            "clear_temp_files",
        )
        self.assertEqual(
            direct_commands.match_direct_command("tmpdelete").operation,
            "clear_temp_files",
        )
        self.assertEqual(
            direct_commands.match_direct_command("WSL圧縮").operation,
            "prune_docker_and_compact_wsl",
        )
        self.assertEqual(
            direct_commands.match_direct_command("wslcomp").operation,
            "prune_docker_and_compact_wsl",
        )
        self.assertEqual(
            direct_commands.match_direct_command("ＷＳＬ圧縮"),
            None,
        )
        self.assertIsNone(
            direct_commands.match_direct_command("tmp削除して"),
        )

    def test_settings_direct_command_parses_allowlisted_page(self) -> None:
        command = direct_commands.match_direct_command(
            "settings windows_update"
        )

        self.assertIsNotNone(command)
        assert command is not None
        self.assertEqual(command.operation, "open_settings")
        self.assertEqual(command.arguments, {"page": "windows_update"})
        with self.assertRaisesRegex(ValueError, "未対応の設定ページ"):
            direct_commands.match_direct_command("settings arbitrary")

    def test_system_info_and_empty_trash_are_direct_commands(self) -> None:
        self.assertEqual(
            direct_commands.match_direct_command("sysinfo").operation,
            "show_system_info",
        )
        self.assertEqual(
            direct_commands.match_direct_command("emptytrash").operation,
            "empty_recycle_bin",
        )

    def test_empty_trash_direct_command_requires_confirmation(self) -> None:
        function = Mock(return_value="ゴミ箱を空にしました。")
        command = direct_commands.DirectCommand(
            "empty_recycle_bin",
            function,
            confirmation_message="実行しますか？",
        )

        result = direct_commands.execute_direct_command(
            command,
            input_function=lambda _: "n",
        )

        self.assertEqual(result, "操作をキャンセルしました。")
        function.assert_not_called()

    @patch("local_actions.cli.select_actions")
    @patch("local_actions.cli.execute_direct_command")
    @patch("local_actions.cli.try_write_action_log")
    def test_direct_command_does_not_call_ollama(
        self,
        write_log: Mock,
        execute_direct: Mock,
        select_actions: Mock,
    ) -> None:
        execute_direct.return_value = "一時フォルダを掃除しました。"

        with (
            patch("local_actions.cli.sys.argv", ["main.py", "tmp削除"]),
            patch("local_actions.cli.print"),
        ):
            cli.main()

        select_actions.assert_not_called()
        execute_direct.assert_called_once()
        write_log.assert_called_once_with(
            "tmp削除",
            "clear_temp_files",
            {},
            "succeeded",
            result="一時フォルダを掃除しました。",
        )

    def test_clear_temp_files_skips_locked_entries(self) -> None:
        with TemporaryDirectory() as directory:
            temp_directory = Path(directory)
            removable = temp_directory / "remove.txt"
            locked = temp_directory / "locked.txt"
            removable.write_text("remove", encoding="utf-8")
            locked.write_text("locked", encoding="utf-8")
            original_unlink = Path.unlink

            def unlink(path: Path, *args: object, **kwargs: object) -> None:
                if path == locked:
                    raise PermissionError("使用中")
                original_unlink(path, *args, **kwargs)

            with (
                patch("local_actions.actions.sys.platform", "win32"),
                patch(
                    "local_actions.actions.tempfile.gettempdir",
                    return_value=directory,
                ),
                patch("pathlib.Path.unlink", new=unlink),
            ):
                result = main.clear_temp_files()

        self.assertIn("削除: 1件", result)
        self.assertIn("スキップ: 1件", result)

    @patch("local_actions.actions.subprocess.run")
    def test_docker_prune_and_compact_wsl_use_fixed_commands(
        self,
        run: Mock,
    ) -> None:
        run.return_value = Mock(returncode=0, stdout="", stderr="")
        written_script = ""
        original_named_temp = main.tempfile.NamedTemporaryFile

        def named_temp(*args: object, **kwargs: object):
            nonlocal written_script
            temporary = original_named_temp(*args, **kwargs)
            original_write = temporary.write

            def write(value: str) -> int:
                nonlocal written_script
                written_script = value
                return original_write(value)

            temporary.write = write
            return temporary

        with (
            patch("local_actions.actions.sys.platform", "win32"),
            patch("pathlib.Path.is_file", return_value=True),
            patch(
                "local_actions.actions.tempfile.NamedTemporaryFile",
                side_effect=named_temp,
            ),
        ):
            result = main.prune_docker_and_compact_wsl()

        self.assertEqual(
            [call.args[0] for call in run.call_args_list[:3]],
            [
                [
                    "wsl.exe",
                    "--exec",
                    "docker",
                    "image",
                    "prune",
                    "--all",
                    "--force",
                ],
                [
                    "wsl.exe",
                    "--exec",
                    "docker",
                    "volume",
                    "prune",
                    "--force",
                ],
                [
                    "wsl.exe",
                    "--exec",
                    "docker",
                    "builder",
                    "prune",
                    "--force",
                ],
            ],
        )
        self.assertEqual(run.call_args_list[3].args[0], ["wsl.exe", "--shutdown"])
        self.assertIn(
            f'select vdisk file="{main.WSL_VIRTUAL_DISK_PATH}"',
            written_script,
        )
        self.assertIn("attach vdisk readonly", written_script)
        self.assertIn("compact vdisk", written_script)
        self.assertIn("detach vdisk", written_script)
        self.assertIn(str(main.WSL_VIRTUAL_DISK_PATH), result)


class ActionLogTests(unittest.TestCase):
    def test_write_action_log_appends_utf8_json_lines(self) -> None:
        with TemporaryDirectory() as directory:
            log_path = Path(directory) / "logs" / "actions.jsonl"
            with patch(
                "local_actions.action_log.get_action_log_path",
                return_value=log_path,
            ):
                action_log.write_action_log(
                    "大分駅の近くを検索して",
                    "google_search",
                    {"query": "大分駅の近く"},
                    "succeeded",
                    result="https://example.com/日本語",
                )
                action_log.write_action_log(
                    "キャンセルして",
                    "empty_recycle_bin",
                    {},
                    "cancelled",
                    result="操作をキャンセルしました。",
                )

            entries = [
                json.loads(line)
                for line in log_path.read_text(encoding="utf-8").splitlines()
            ]

        self.assertEqual(len(entries), 2)
        self.assertEqual(entries[0]["request"], "大分駅の近くを検索して")
        self.assertEqual(entries[0]["arguments"], {"query": "大分駅の近く"})
        self.assertEqual(entries[0]["status"], "succeeded")
        self.assertIsNone(entries[0]["error"])
        self.assertEqual(entries[1]["status"], "cancelled")
        self.assertIn("T", entries[0]["timestamp"])

    def test_log_path_uses_onedrive_directory(self) -> None:
        onedrive_directory = Path(r"C:\Users\test\OneDrive\Local Actions")
        with patch(
            "local_actions.action_log.get_onedrive_directory",
            return_value=onedrive_directory,
        ):
            self.assertEqual(
                action_log.get_action_log_path(),
                onedrive_directory / "actions.jsonl",
            )

    @patch("local_actions.cli.execute_workflow")
    @patch("local_actions.cli.select_actions")
    @patch("local_actions.cli.try_write_action_log")
    def test_main_records_success(
        self,
        write_log: Mock,
        select_actions: Mock,
        execute_workflow: Mock,
    ) -> None:
        select_actions.return_value = [
            registry.PlannedAction("google_maps_search", {"query": "大分駅"})
        ]
        execute_workflow.return_value = registry.ActionExecutionResult(
            "succeeded",
            "https://www.google.com/maps/search/?api=1&query=大分駅",
        )

        with patch("local_actions.cli.sys.argv", ["main.py", "大分駅を地図で検索"]):
            cli.main()

        write_log.assert_called_once_with(
            "大分駅を地図で検索",
            "google_maps_search",
            {"query": "大分駅"},
            "succeeded",
            result="https://www.google.com/maps/search/?api=1&query=大分駅",
        )

    @patch(
        "local_actions.cli.select_actions",
        side_effect=RuntimeError("Ollama停止"),
    )
    @patch("local_actions.cli.try_write_action_log")
    def test_main_records_selection_failure(
        self,
        write_log: Mock,
        select_action: Mock,
    ) -> None:
        with (
            patch("local_actions.cli.sys.argv", ["main.py", "検索して"]),
            self.assertRaisesRegex(RuntimeError, "Ollama停止"),
        ):
            cli.main()

        write_log.assert_called_once_with(
            "検索して",
            None,
            {},
            "failed",
            error="RuntimeError: Ollama停止",
        )


class RecycleBinActionTests(unittest.TestCase):
    @patch("local_actions.actions.os.startfile")
    def test_open_recycle_bin_uses_shell_folder(self, startfile: Mock) -> None:
        self.assertEqual(main.open_recycle_bin(), "ゴミ箱を開きました。")
        startfile.assert_called_once_with("shell:RecycleBinFolder")

    def test_action_with_arguments_rejects_unknown_field(self) -> None:
        with self.assertRaisesRegex(ValueError, "未定義の引数"):
            registry.execute_action(
                "google_maps_search",
                {"query": "生成AI", "args": {}},
            )

    @patch("local_actions.actions.subprocess.run")
    def test_empty_recycle_bin_uses_fixed_powershell_command(
        self,
        run: Mock,
    ) -> None:
        run.return_value = Mock(stderr="")

        with patch("local_actions.actions.sys.platform", "win32"):
            self.assertEqual(main.empty_recycle_bin(), "ゴミ箱を空にしました。")

        run.assert_called_once_with(
            [
                "powershell",
                "-NoProfile",
                "-Command",
                "Clear-RecycleBin -Force -ErrorAction SilentlyContinue",
            ],
            capture_output=True,
            text=True,
            encoding="utf-8",
        )

    @patch("local_actions.actions.subprocess.run")
    @patch("local_actions.actions.os.startfile")
    def test_copilot_chat_opens_allowed_url_and_passes_message_env(
        self,
        startfile: Mock,
        run: Mock,
    ) -> None:
        run.return_value = Mock(returncode=0, stdout="", stderr="")

        with patch("local_actions.actions.sys.platform", "win32"):
            result = main.copilot_chat("要約して", "Opus")

        startfile.assert_called_once_with(main.COPILOT_URL)
        command = run.call_args.args[0]
        environment = run.call_args.kwargs["env"]
        self.assertEqual(command[:5], [
            "powershell.exe",
            "-NoProfile",
            "-NonInteractive",
            "-Sta",
            "-Command",
        ])
        self.assertEqual(environment["LOCAL_ACTIONS_COPILOT_MESSAGE"], "要約して")
        self.assertEqual(environment["LOCAL_ACTIONS_COPILOT_MODEL"], "Opus")
        self.assertIn("Copilot", result)

class UiaAgentTests(unittest.TestCase):
    def test_validate_uia_plan_accepts_known_element(self) -> None:
        elements = [
            uia_agent.UiaElement(
                id="e0",
                index=0,
                type="Button",
                name="OK",
                bounds={"x": 1, "y": 2, "width": 30, "height": 20},
                enabled=True,
            )
        ]

        plan = uia_agent.validate_uia_plan(
            {"element_id": "e0", "action": "click", "reason": "matches"},
            elements,
        )

        self.assertEqual(plan.element_id, "e0")
        self.assertEqual(plan.action, "click")


    def test_validate_uia_plan_clears_text_for_click(self) -> None:
        elements = [
            uia_agent.UiaElement(
                id="e0",
                index=0,
                type="Button",
                name="OK",
                bounds={"x": 1, "y": 2, "width": 30, "height": 20},
                enabled=True,
            )
        ]

        plan = uia_agent.validate_uia_plan(
            {"element_id": "e0", "action": "click", "text": "ignored"},
            elements,
        )

        self.assertEqual(plan.text, "")

    def test_candidate_elements_prefers_request_matching_names(self) -> None:
        elements = [
            uia_agent.UiaElement(
                id="e0",
                index=0,
                type="Button",
                name="こんにちはの返答迷走 の会話オプションを開く",
                bounds={"x": 1, "y": 2, "width": 30, "height": 20},
                enabled=True,
            ),
            uia_agent.UiaElement(
                id="e1",
                index=1,
                type="Button",
                name="送信",
                bounds={"x": 1, "y": 2, "width": 30, "height": 20},
                enabled=True,
            ),
        ]

        candidates = uia_agent.candidate_elements_for_request("送信して", elements)

        self.assertEqual(candidates[0].id, "e1")

    def test_validate_uia_plan_rejects_unknown_element(self) -> None:
        elements = [
            uia_agent.UiaElement(
                id="e0",
                index=0,
                type="Button",
                name="OK",
                bounds={"x": 1, "y": 2, "width": 30, "height": 20},
                enabled=True,
            )
        ]

        with self.assertRaisesRegex(ValueError, "Unknown element_id"):
            uia_agent.validate_uia_plan(
                {"element_id": "e1", "action": "click"},
                elements,
            )

    def test_select_uia_plan_with_tools_accepts_single_tool_call(self) -> None:
        elements = [
            uia_agent.UiaElement(
                id="e0",
                index=0,
                type="Button",
                name="OK",
                bounds={"x": 1, "y": 2, "width": 30, "height": 20},
                enabled=True,
            )
        ]
        response = Mock(
            message=Mock(
                content="",
                tool_calls=[
                    SimpleNamespace(function=SimpleNamespace(name="click", arguments={"element_id": "e0"}))
                ],
            )
        )

        with patch("local_actions.uia_agent.chat", return_value=response) as chat:
            plan = uia_agent.select_uia_plan_with_tools("OK を押して", elements)

        self.assertEqual(plan.element_id, "e0")
        self.assertEqual(plan.action, "click")
        self.assertIn("tools", chat.call_args.kwargs)
        self.assertNotIn("format", chat.call_args.kwargs)

    def test_tool_call_payload_rejects_missing_tool_call_with_content(self) -> None:
        response = Mock(
            message=Mock(
                content="JSON を返してしまいました",
                tool_calls=[],
            )
        )

        with self.assertRaisesRegex(ValueError, "JSON を返してしまいました"):
            uia_agent._single_tool_call_payload(response, {"click"})

    def test_tool_call_payload_rejects_multiple_tool_calls(self) -> None:
        response = Mock(
            message=Mock(
                content="",
                tool_calls=[
                    SimpleNamespace(function=SimpleNamespace(name="click", arguments={"element_id": "e0"})),
                    SimpleNamespace(function=SimpleNamespace(name="focus", arguments={"element_id": "e1"})),
                ],
            )
        )

        with self.assertRaisesRegex(ValueError, "multiple UI Automation tool calls"):
            uia_agent._single_tool_call_payload(response, {"click", "focus"})

    def test_select_uia_loop_decision_with_tools_accepts_wait(self) -> None:
        elements = [
            uia_agent.UiaElement(
                id="e0",
                index=0,
                type="Button",
                name="OK",
                bounds={"x": 1, "y": 2, "width": 30, "height": 20},
                enabled=True,
            )
        ]
        response = Mock(
            message=Mock(
                content="",
                tool_calls=[SimpleNamespace(function=SimpleNamespace(name="wait", arguments={"seconds": 9}))],
            )
        )

        with patch("local_actions.uia_agent.chat", return_value=response):
            decision = uia_agent.select_uia_loop_decision_with_tools(
                request="待って",
                elements=elements,
                diff=uia_agent.diff_snapshots(None, elements),
                last_action=None,
                turn=1,
                max_turns=3,
            )

        self.assertEqual(decision.action, "wait")
        self.assertEqual(decision.seconds, 5.0)

    def test_select_uia_loop_decision_with_tools_accepts_done(self) -> None:
        elements = [
            uia_agent.UiaElement(
                id="e0",
                index=0,
                type="Button",
                name="OK",
                bounds={"x": 1, "y": 2, "width": 30, "height": 20},
                enabled=True,
            )
        ]
        response = Mock(
            message=Mock(
                content="",
                tool_calls=[SimpleNamespace(function=SimpleNamespace(name="done", arguments={}))],
            )
        )

        with patch("local_actions.uia_agent.chat", return_value=response):
            decision = uia_agent.select_uia_loop_decision_with_tools(
                request="終わって",
                elements=elements,
                diff=uia_agent.diff_snapshots(None, elements),
                last_action=None,
                turn=1,
                max_turns=3,
            )

        self.assertEqual(decision.action, "done")

    def test_select_uia_plan_with_tools_rejects_disabled_element(self) -> None:
        elements = [
            uia_agent.UiaElement(
                id="e0",
                index=0,
                type="Button",
                name="OK",
                bounds={"x": 1, "y": 2, "width": 30, "height": 20},
                enabled=False,
            )
        ]
        response = Mock(
            message=Mock(
                content="",
                tool_calls=[
                    SimpleNamespace(function=SimpleNamespace(name="click", arguments={"element_id": "e0"}))
                ],
            )
        )

        with (
            patch("local_actions.uia_agent.chat", return_value=response),
            self.assertRaisesRegex(ValueError, "disabled"),
        ):
            uia_agent.select_uia_plan_with_tools("OK を押して", elements)

    def test_list_window_elements_parses_powershell_json(self) -> None:
        payload = {
            "window": {"name": "Calculator"},
            "elements": [
                {
                    "id": "e0",
                    "index": 0,
                    "automation_id": "num1Button",
                    "type": "Button",
                    "name": "One",
                    "bounds": {"x": 10, "y": 20, "width": 30, "height": 40},
                    "enabled": True,
                }
            ],
        }

        with (
            patch("local_actions.uia_agent.sys.platform", "win32"),
            patch("local_actions.uia_agent.subprocess.run") as run,
        ):
            run.return_value = Mock(
                returncode=0,
                stdout=json.dumps(payload),
                stderr="",
            )
            elements = uia_agent.list_window_elements()

        self.assertEqual(elements[0].id, "e0")
        self.assertEqual(elements[0].automation_id, "num1Button")
        self.assertEqual(elements[0].bounds["width"], 30)
        self.assertEqual(run.call_args.args[0][:5], [
            "powershell.exe",
            "-NoProfile",
            "-NonInteractive",
            "-Sta",
            "-Command",
        ])

    def test_execute_uia_plan_rejects_disabled_element_before_powershell(self) -> None:
        elements = [
            uia_agent.UiaElement(
                id="e0",
                index=0,
                type="Button",
                name="OK",
                bounds={"x": 1, "y": 2, "width": 30, "height": 20},
                enabled=False,
            )
        ]

        with self.assertRaisesRegex(ValueError, "disabled"):
            uia_agent.execute_uia_plan(
                uia_agent.UiaPlan(element_id="e0", action="click"),
                elements,
            )

    def test_main_list_elements_prints_elements_without_slm(self) -> None:
        elements = [
            uia_agent.UiaElement(
                id="e0",
                index=0,
                type="Button",
                name="OK",
                bounds={"x": 1, "y": 2, "width": 30, "height": 20},
                enabled=True,
                automation_id="okButton",
            )
        ]

        with (
            patch("local_actions.uia_agent.sys.argv", ["uia_agent.py", "--list-elements"]),
            patch("local_actions.uia_agent.list_window_elements", return_value=elements) as list_elements,
            patch("local_actions.uia_agent.chat") as chat,
            patch("builtins.print") as print_mock,
        ):
            uia_agent.main()

        list_elements.assert_called_once_with(window_title="", foreground=True)
        chat.assert_not_called()
        printed = print_mock.call_args.args[0]
        payload = json.loads(printed)
        self.assertEqual(payload[0]["id"], "e0")
        self.assertEqual(payload[0]["automation_id"], "okButton")

if __name__ == "__main__":
    unittest.main()
