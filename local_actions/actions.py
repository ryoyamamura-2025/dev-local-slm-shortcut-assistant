import ctypes
import json
import os
import platform
import shutil
import subprocess
import sys
import tempfile
import time
import uuid
from collections.abc import Callable
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Literal, get_args
from urllib.parse import quote, urlencode

from openpyxl import Workbook, load_workbook


CLEANUP_LIST_PATH = Path(__file__).with_name("pending_cleanup.txt")
COPILOT_URL = "https://m365.cloud.microsoft/chat"

COPILOT_CHAT_SCRIPT = r"""
$ErrorActionPreference = "Stop"
[Console]::OutputEncoding = [System.Text.UTF8Encoding]::new($false)

$Message = $env:LOCAL_ACTIONS_COPILOT_MESSAGE
$Model   = $env:LOCAL_ACTIONS_COPILOT_MODEL

Add-Type -AssemblyName UIAutomationClient
Add-Type -AssemblyName UIAutomationTypes
Add-Type -AssemblyName System.Windows.Forms

Add-Type -TypeDefinition @"
using System;
using System.Runtime.InteropServices;
using System.Text;
public static class Win32 {
    public delegate bool EnumChildProc(IntPtr hWnd, IntPtr lParam);
    [DllImport("user32.dll")] public static extern bool EnumChildWindows(IntPtr p, EnumChildProc cb, IntPtr lp);
    [DllImport("user32.dll")] public static extern bool SetForegroundWindow(IntPtr hWnd);
    [DllImport("user32.dll")] public static extern bool ShowWindow(IntPtr hWnd, int n);
    [DllImport("user32.dll")] public static extern bool IsIconic(IntPtr hWnd);
    [DllImport("user32.dll")] public static extern bool SetCursorPos(int x, int y);
    [DllImport("user32.dll")] public static extern void mouse_event(uint f, int dx, int dy, uint c, uint e);
    [DllImport("user32.dll", CharSet=CharSet.Unicode)]
    public static extern int GetClassName(IntPtr hWnd, StringBuilder sb, int max);

    public static IntPtr FindChildByClass(IntPtr parent, string cls) {
        IntPtr found = IntPtr.Zero;
        EnumChildProc cb = (hWnd, _) => {
            var sb = new StringBuilder(256);
            GetClassName(hWnd, sb, 256);
            if (sb.ToString() == cls) { found = hWnd; return false; }
            return true;
        };
        EnumChildWindows(parent, cb, IntPtr.Zero);
        GC.KeepAlive(cb);
        return found;
    }
}
"@

function Click-Element {
    param([System.Windows.Automation.AutomationElement]$El)
    # Try InvokePattern, then mouse click, then walk up to parent (child text elements have no clickable point)
    $walker = [System.Windows.Automation.TreeWalker]::ControlViewWalker
    $target = $El
    for ($i = 0; $i -lt 4; $i++) {
        try {
            $target.GetCurrentPattern([System.Windows.Automation.InvokePattern]::Pattern).Invoke()
            return
        } catch {}
        try {
            $pt = $target.GetClickablePoint()
            [Win32]::SetCursorPos([int]$pt.X, [int]$pt.Y) | Out-Null
            Start-Sleep -Milliseconds 80
            [Win32]::mouse_event(0x0002, [int]$pt.X, [int]$pt.Y, 0, 0)
            Start-Sleep -Milliseconds 50
            [Win32]::mouse_event(0x0004, [int]$pt.X, [int]$pt.Y, 0, 0)
            return
        } catch {}
        $parent = $walker.GetParent($target)
        if ($null -eq $parent) { break }
        $target = $parent
    }
    throw "要素をクリックできませんでした"
}

# Search one control type per call (avoids nested ::new parsing issues)
function Find-ByName {
    param(
        [System.Windows.Automation.AutomationElement]$Root,
        [System.Windows.Automation.ControlType]$Type,
        [string]$NamePattern,
        [int]$Sec = 15
    )
    $typeCond = [System.Windows.Automation.PropertyCondition]::new(
        [System.Windows.Automation.AutomationElement]::ControlTypeProperty,
        $Type
    )
    $until = (Get-Date).AddSeconds($Sec)
    while ((Get-Date) -lt $until) {
        $els = $Root.FindAll([System.Windows.Automation.TreeScope]::Descendants, $typeCond)
        foreach ($el in $els) {
            try { if ($el.Current.Name -like $NamePattern) { return $el } } catch {}
        }
        Start-Sleep -Milliseconds 300
    }
    return $null
}

# Search multiple control types in each polling iteration (no sequential timeout waste)
function Find-ByNameAny {
    param(
        [System.Windows.Automation.AutomationElement]$Root,
        [System.Windows.Automation.ControlType[]]$Types,
        [string]$NamePattern,
        [int]$Sec = 8
    )
    $until = (Get-Date).AddSeconds($Sec)
    while ((Get-Date) -lt $until) {
        foreach ($ct in $Types) {
            $typeCond = [System.Windows.Automation.PropertyCondition]::new(
                [System.Windows.Automation.AutomationElement]::ControlTypeProperty,
                $ct
            )
            $els = $Root.FindAll([System.Windows.Automation.TreeScope]::Descendants, $typeCond)
            foreach ($el in $els) {
                try {
                    if ($el.Current.Name -like $NamePattern -and -not $el.Current.IsOffscreen) {
                        return $el
                    }
                } catch {}
            }
        }
        Start-Sleep -Milliseconds 300
    }
    return $null
}

# 1. Find Edge window with Copilot tab
$edgeWindow = $null
$until = (Get-Date).AddSeconds(15)
while ((Get-Date) -lt $until) {
    $wins = ([System.Windows.Automation.AutomationElement]::RootElement).FindAll(
        [System.Windows.Automation.TreeScope]::Children,
        [System.Windows.Automation.Condition]::TrueCondition
    )
    foreach ($w in $wins) {
        try {
            if ($w.Current.ClassName -like "*Chrome_WidgetWin*" -and
                $w.Current.Name -like "*Copilot*") {
                $edgeWindow = $w; break
            }
        } catch {}
    }
    if ($null -ne $edgeWindow) { break }
    Start-Sleep -Milliseconds 500
}
if ($null -eq $edgeWindow) { throw "M365 CopilotのEdgeウィンドウが見つかりません。" }

# 2. Restore only if minimized to avoid un-maximizing, then bring to front
$hwnd = [IntPtr]::new($edgeWindow.Current.NativeWindowHandle)
if ([Win32]::IsIconic($hwnd)) { [Win32]::ShowWindow($hwnd, 9) | Out-Null }
[Win32]::SetForegroundWindow($hwnd) | Out-Null
Start-Sleep -Milliseconds 600

# 3. Find render widget via Win32 child enumeration, fall back to edge window
$renderHwnd = [Win32]::FindChildByClass($hwnd, "Chrome_RenderWidgetHostHWND")
if ($renderHwnd -ne [IntPtr]::Zero) {
    try { $searchRoot = [System.Windows.Automation.AutomationElement]::FromHandle($renderHwnd) }
    catch { $searchRoot = $edgeWindow }
} else { $searchRoot = $edgeWindow }

# 4. Change model if needed
if ($Model -ne "自動") {
    # Try Button / SplitButton / ComboBox across render widget then full edge window
    $btn = $null
    :btnSearch foreach ($root in @($searchRoot, $edgeWindow)) {
        foreach ($ct in @(
            [System.Windows.Automation.ControlType]::Button,
            [System.Windows.Automation.ControlType]::SplitButton,
            [System.Windows.Automation.ControlType]::ComboBox
        )) {
            $btn = Find-ByName -Root $root -Type $ct -NamePattern "*モデル セレクター*" -Sec 3
            if ($null -ne $btn) { break btnSearch }
        }
    }

    if ($null -eq $btn) {
        # Diagnostic: dump all Button names to identify actual element name
        $diagCond = [System.Windows.Automation.PropertyCondition]::new(
            [System.Windows.Automation.AutomationElement]::ControlTypeProperty,
            [System.Windows.Automation.ControlType]::Button
        )
        $diagEls = $edgeWindow.FindAll([System.Windows.Automation.TreeScope]::Descendants, $diagCond)
        $diagNames = @()
        foreach ($el in $diagEls) {
            try { $n = $el.Current.Name; if ($n) { $diagNames += $n } } catch {}
        }
        throw "モデル選択ボタン（自動）が見つかりません。`n検出されたButton名: $(($diagNames | Sort-Object -Unique) -join ' | ')"
    }

    Click-Element $btn
    Start-Sleep -Milliseconds 500  # Wait for Fluent UI popup to render

    # TrueCondition search — finds Opus regardless of control type or offscreen state
    $opt = $null
    $until = (Get-Date).AddSeconds(10)
    while ((Get-Date) -lt $until) {
        $allEls = $edgeWindow.FindAll(
            [System.Windows.Automation.TreeScope]::Descendants,
            [System.Windows.Automation.Condition]::TrueCondition
        )
        foreach ($el in $allEls) {
            try {
                if ($el.Current.Name -like "*$Model*" -and $el.Current.IsEnabled -and
                    $el.Current.Name -ne "モデル セレクター") {
                    $opt = $el; break
                }
            } catch {}
        }
        if ($null -ne $opt) { break }
        Start-Sleep -Milliseconds 400
    }

    if ($null -eq $opt) { throw "${Model}オプションが見つかりません。" }
    Click-Element $opt
    Start-Sleep -Milliseconds 500
}

# 5. Find chat input: Edit (any name), then focusable Document
$chatInput = $null
:chatSearch foreach ($root in @($searchRoot, $edgeWindow)) {
    $chatInput = Find-ByName -Root $root -Type ([System.Windows.Automation.ControlType]::Edit) -NamePattern "*" -Sec 5
    if ($null -ne $chatInput) { break chatSearch }
    $docTypeCond = [System.Windows.Automation.PropertyCondition]::new(
        [System.Windows.Automation.AutomationElement]::ControlTypeProperty,
        [System.Windows.Automation.ControlType]::Document
    )
    $focusCond = [System.Windows.Automation.PropertyCondition]::new(
        [System.Windows.Automation.AutomationElement]::IsKeyboardFocusableProperty,
        $true
    )
    $docCond = [System.Windows.Automation.AndCondition]::new($docTypeCond, $focusCond)
    $chatInput = $root.FindFirst([System.Windows.Automation.TreeScope]::Descendants, $docCond)
    if ($null -ne $chatInput) { break chatSearch }
}
if ($null -eq $chatInput) {
    # Diagnostic: dump Edit and Document names
    $diagNames = @()
    foreach ($ct in @([System.Windows.Automation.ControlType]::Edit, [System.Windows.Automation.ControlType]::Document)) {
        $cond = [System.Windows.Automation.PropertyCondition]::new(
            [System.Windows.Automation.AutomationElement]::ControlTypeProperty, $ct
        )
        $els = $edgeWindow.FindAll([System.Windows.Automation.TreeScope]::Descendants, $cond)
        foreach ($el in $els) {
            try { $diagNames += "$($ct.ProgrammaticName): $($el.Current.Name)" } catch {}
        }
    }
    throw "チャット入力欄が見つかりません。`n検出されたEdit/Document名: $($diagNames -join ' | ')"
}

# 6. Paste and send
$chatInput.SetFocus()
Start-Sleep -Milliseconds 200
[System.Windows.Forms.Clipboard]::SetText($Message)
[System.Windows.Forms.SendKeys]::SendWait("^v")
Start-Sleep -Milliseconds 300
[System.Windows.Forms.SendKeys]::SendWait("{ENTER}")

"Copilot（モデル: $Model）にメッセージを送信しました。"
"""
WSL_VIRTUAL_DISK_PATH = Path(
    r"C:\Users\k250501260\AppData\Local\Packages"
    r"\CanonicalGroupLimited.Ubuntu22.04LTS_79rhkp1fndgsc\LocalState\ext4.vhdx"
)

ELEVATED_DISKPART_SCRIPT = r"""
$ErrorActionPreference = "Stop"
$argument = '/s "{0}"' -f $env:LOCAL_ACTIONS_DISKPART_SCRIPT
$process = Start-Process `
    -FilePath "$env:SystemRoot\System32\diskpart.exe" `
    -ArgumentList $argument `
    -Verb RunAs `
    -WindowStyle Hidden `
    -Wait `
    -PassThru
exit $process.ExitCode
"""

SettingsPage = Literal[
    "system",
    "display",
    "sound",
    "notifications",
    "network",
    "bluetooth",
    "apps",
    "default_apps",
    "storage",
    "power",
    "privacy",
    "windows_update",
]

FolderName = Literal["onedrive", "downloads"]

CalendarEntityType = Literal["タスク", "リマインダ"]
CaptureKind = Literal["memo", "log"]
XDestination = Literal["search", "home", "profile", "likes"]
X_USERNAME_ENV_VAR = "LOCAL_ACTIONS_X_USERNAME"

CALENDAR_ENTITY_TYPES: tuple[str, ...] = get_args(CalendarEntityType)
CAPTURE_KINDS: tuple[str, ...] = get_args(CaptureKind)
CAPTURE_HEADERS = ("timestamp", "kind", "title", "body")

SETTINGS_PAGES: dict[str, tuple[str, str]] = {
    "system": ("システム情報", "ms-settings:about"),
    "display": ("ディスプレイ", "ms-settings:display"),
    "sound": ("サウンド", "ms-settings:sound"),
    "notifications": ("通知", "ms-settings:notifications"),
    "network": ("ネットワークとインターネット", "ms-settings:network-status"),
    "bluetooth": ("Bluetoothとデバイス", "ms-settings:bluetooth"),
    "apps": ("インストールされているアプリ", "ms-settings:appsfeatures"),
    "default_apps": ("既定のアプリ", "ms-settings:defaultapps"),
    "storage": ("ストレージ", "ms-settings:storagesense"),
    "power": ("電源", "ms-settings:powersleep"),
    "privacy": ("プライバシーとセキュリティ", "ms-settings:privacy"),
    "windows_update": ("Windows Update", "ms-settings:windowsupdate"),
}

COPY_TEXT_SCRIPT = r"""
$ErrorActionPreference = "Stop"
[Console]::InputEncoding = [System.Text.UTF8Encoding]::new($false)
Add-Type -AssemblyName System.Windows.Forms
$text = [Console]::In.ReadToEnd()
[System.Windows.Forms.Clipboard]::SetText($text)
"""

GET_CLIPBOARD_TEXT_SCRIPT = r"""
$ErrorActionPreference = "Stop"
[Console]::OutputEncoding = [System.Text.UTF8Encoding]::new($false)
Add-Type -AssemblyName System.Windows.Forms
if (-not [System.Windows.Forms.Clipboard]::ContainsText()) {
    throw "The clipboard does not contain text."
}
[Console]::Out.Write([System.Windows.Forms.Clipboard]::GetText())
"""

CURRENT_PAGE_SCRIPT = r"""
$ErrorActionPreference = "Stop"
[Console]::OutputEncoding = [System.Text.UTF8Encoding]::new($false)

Add-Type -AssemblyName UIAutomationClient
Add-Type -AssemblyName UIAutomationTypes

Add-Type -TypeDefinition @"
using System;
using System.Collections.Generic;
using System.Diagnostics;
using System.Runtime.InteropServices;
using System.Text;

public static class BrowserWindowFinder
{
    private delegate bool EnumWindowsProc(IntPtr hWnd, IntPtr lParam);

    [DllImport("user32.dll")]
    private static extern bool EnumWindows(EnumWindowsProc callback, IntPtr lParam);

    [DllImport("user32.dll")]
    private static extern bool IsWindowVisible(IntPtr hWnd);

    [DllImport("user32.dll", CharSet = CharSet.Unicode)]
    private static extern int GetClassName(
        IntPtr hWnd,
        StringBuilder className,
        int maxCount
    );

    [DllImport("user32.dll")]
    private static extern uint GetWindowThreadProcessId(
        IntPtr hWnd,
        out uint processId
    );

    public static IntPtr[] GetWindows()
    {
        var windows = new List<IntPtr>();

        EnumWindows(delegate (IntPtr hWnd, IntPtr lParam)
        {
            if (!IsWindowVisible(hWnd))
            {
                return true;
            }

            var className = new StringBuilder(256);
            GetClassName(hWnd, className, className.Capacity);
            if (!className.ToString().StartsWith("Chrome_WidgetWin"))
            {
                return true;
            }

            uint processId;
            GetWindowThreadProcessId(hWnd, out processId);

            try
            {
                string processName = Process.GetProcessById((int)processId)
                    .ProcessName.ToLowerInvariant();
                if (processName == "chrome" || processName == "msedge")
                {
                    windows.Add(hWnd);
                }
            }
            catch
            {
            }

            return true;
        }, IntPtr.Zero);

        return windows.ToArray();
    }
}
"@

$browser = $null
foreach ($handle in [BrowserWindowFinder]::GetWindows()) {
    try {
        $candidate = [System.Windows.Automation.AutomationElement]::FromHandle(
            $handle
        )
        if (-not [string]::IsNullOrWhiteSpace($candidate.Current.Name)) {
            $browser = $candidate
            break
        }
    } catch {
    }
}

if ($null -eq $browser) {
    throw "Chrome or Edge window was not found."
}

$editCondition = [System.Windows.Automation.PropertyCondition]::new(
    [System.Windows.Automation.AutomationElement]::ControlTypeProperty,
    [System.Windows.Automation.ControlType]::Edit
)
$edits = $browser.FindAll(
    [System.Windows.Automation.TreeScope]::Descendants,
    $editCondition
)
$candidates = @()

foreach ($edit in $edits) {
    try {
        $name = $edit.Current.Name
        $valuePattern = $edit.GetCurrentPattern(
            [System.Windows.Automation.ValuePattern]::Pattern
        )
        $value = $valuePattern.Current.Value

        if ([string]::IsNullOrWhiteSpace($value)) {
            continue
        }

        $isUrl = (
            $value -match "^[a-zA-Z][a-zA-Z0-9+.-]*:" -or
            $value -match "^[^\s]+\.[^\s]+"
        )
        if (-not $isUrl) {
            continue
        }

        $score = 0
        if ($name -match "(?i)address|location|url|アドレス") {
            $score += 100
        }
        if ($value -match "^[a-zA-Z][a-zA-Z0-9+.-]*:") {
            $score += 50
        } else {
            $score += 10
        }

        if ($score -gt 0) {
            $candidates += [PSCustomObject]@{
                Value = $value
                Score = $score
            }
        }
    } catch {
    }
}

$address = $candidates |
    Sort-Object Score -Descending |
    Select-Object -First 1

if ($null -eq $address) {
    throw "Browser address bar was not found."
}

$title = $browser.Current.Name -replace (
    "\s+[-–]\s+(Google Chrome|Microsoft Edge)$"
), ""

[PSCustomObject]@{
    title = $title
    url = $address.Value
} | ConvertTo-Json -Compress
"""


def google_search(query: str) -> str:
    """Googleで検索する。

    Args:
        query: 日本語を保持した検索語。
    """
    return "https://www.google.com/search?" + urlencode({"q": query})


def google_maps_search(query: str) -> str:
    """Googleマップで場所や店舗を検索する。

    Args:
        query: 日本語を保持した検索語。
    """
    return "https://www.google.com/maps/search/?" + urlencode(
        {"api": "1", "query": query}
    )


def get_x_username() -> str:
    """環境変数に固定されたXのユーザー名を検証して返す。"""
    username = os.environ.get(X_USERNAME_ENV_VAR, "").strip().removeprefix("@")
    if not username:
        raise OSError(
            f"環境変数{X_USERNAME_ENV_VAR}にXのユーザー名を設定してください。"
        )
    if (
        len(username) > 15
        or any(
            character
            not in "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789_"
            for character in username
        )
    ):
        raise ValueError(
            f"環境変数{X_USERNAME_ENV_VAR}のXユーザー名が不正です。"
        )
    return username


def x_open(destination: XDestination, query: str = "") -> str:
    """Xの検索、ホーム、プロフィール、いいねを開く。

    Args:
        destination: 開く場所。search、home、profile、likesのいずれか。
        query: searchで使う検索語。その他では空文字。
    """
    if destination == "search":
        if not query.strip():
            raise ValueError("Xの検索語が空です。")
        return "https://x.com/search?" + urlencode(
            {"q": query, "src": "typed_query"}
        )
    if destination == "home":
        return "https://x.com/home"
    if destination in {"profile", "likes"}:
        username = get_x_username()
        suffix = "/likes" if destination == "likes" else ""
        return f"https://x.com/{quote(username, safe='')}{suffix}"
    raise ValueError(f"未対応のX表示先です: {destination}")


def get_onedrive_directory() -> Path:
    """現在のユーザーのOneDrive内にあるアプリ用フォルダを返す。"""
    if sys.platform != "win32":
        raise OSError("OneDriveへの保存はWindowsでのみ利用できます。")

    for variable in ("OneDriveConsumer", "OneDrive", "OneDriveCommercial"):
        value = os.environ.get(variable)
        if value and Path(value).is_dir():
            return Path(value) / "Local Actions"

    raise OSError(
        "OneDriveフォルダを特定できませんでした。"
        "OneDriveへサインインしてから再実行してください。"
    )


def get_capture_path() -> Path:
    """キャプチャを書き込むOneDrive上のExcelファイルを返す。"""
    return get_onedrive_directory() / "inbox.xlsx"


def get_downloads_directory() -> Path:
    """現在のユーザーのダウンロードフォルダを返す。"""
    return Path.home() / "Downloads"


FOLDERS: dict[str, tuple[str, Callable[[], Path]]] = {
    "onedrive": ("OneDriveの保存先", get_onedrive_directory),
    "downloads": ("ダウンロードフォルダ", get_downloads_directory),
}


def open_folder(folder: FolderName) -> str:
    """許可済みのフォルダをExplorerで開く。

    Args:
        folder: 開くフォルダ。onedriveまたはdownloads。
    """
    if sys.platform != "win32":
        raise OSError("フォルダを開く操作はWindowsでのみ利用できます。")

    target = FOLDERS.get(folder)
    if target is None:
        allowed = ", ".join(FOLDERS)
        raise ValueError(
            f"未対応のフォルダです: {folder}（許可値: {allowed}）"
        )

    label, resolver = target
    path = resolver()
    if not path.is_dir():
        raise OSError(f"{label}が見つかりません: {path}")
    os.startfile(path)
    return f"{label}を開きました。\n{path}"


def open_settings(page: SettingsPage) -> str:
    """許可済みのWindows設定ページを開く。

    Args:
        page: 開くページ。system、display、sound、notifications、network、
            bluetooth、apps、default_apps、storage、power、privacy、
            windows_updateのいずれか。
    """
    if sys.platform != "win32":
        raise OSError("Windows設定を開く操作はWindowsでのみ利用できます。")

    setting = SETTINGS_PAGES.get(page)
    if setting is None:
        allowed = ", ".join(SETTINGS_PAGES)
        raise ValueError(f"未対応の設定ページです: {page}（許可値: {allowed}）")

    label, uri = setting
    os.startfile(uri)
    return f"Windows設定の「{label}」を開きました。"


def run_clipboard_script(script: str, text: str | None = None) -> str:
    """固定PowerShellスクリプトでWindowsクリップボードを操作する。

    Args:
        script: Python側で定義した固定スクリプト。
        text: スクリプトへUTF-8で渡す標準入力。
    """
    if sys.platform != "win32":
        raise OSError("クリップボード操作はWindowsでのみ利用できます。")

    completed = subprocess.run(
        [
            "powershell.exe",
            "-NoProfile",
            "-NonInteractive",
            "-Sta",
            "-Command",
            script,
        ],
        input=text,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        check=False,
    )
    if completed.returncode != 0:
        detail = completed.stderr.strip()
        raise OSError(
            "クリップボードを操作できませんでした。"
            + (f"\n{detail}" if detail else "")
        )
    return completed.stdout


def copy_text(text: str) -> str:
    """指定されたテキストを改変せずクリップボードへコピーする。

    Args:
        text: コピーするテキスト。
    """
    if not text:
        raise ValueError("コピーするテキストが空です。")
    run_clipboard_script(COPY_TEXT_SCRIPT, text)
    return f"テキストをクリップボードへコピーしました（{len(text)}文字）。"


def get_clipboard_text() -> str:
    """クリップボードにあるテキストを取得する。"""
    text = run_clipboard_script(GET_CLIPBOARD_TEXT_SCRIPT)
    if not text:
        raise ValueError("クリップボードのテキストが空です。")
    return text


def _normalize_title(title: str, label: str) -> str:
    """必須タイトルから前後の空白を除去し、空ならエラーにする。"""
    normalized_title = title.strip()
    if not normalized_title:
        raise ValueError(f"{label}が空です。")
    return normalized_title


def capture(kind: CaptureKind, title: str, body: str = "") -> str:
    """メモまたはログをOneDrive上のExcel INBOXへ追記する。

    Args:
        kind: キャプチャの種別。memoまたはlog。
        title: キャプチャのタイトル。
        body: キャプチャの本文。依頼文で指定された本文、またはワークフローで
            直前の操作結果が渡る。両方ある場合は結合済みの文字列が渡る。
    """
    normalized_title = _normalize_title(title, "タイトル")
    if kind not in CAPTURE_KINDS:
        allowed = ", ".join(CAPTURE_KINDS)
        raise ValueError(
            f"未対応のキャプチャ種別です: {kind}（許可値: {allowed}）"
        )

    capture_path = get_capture_path()
    capture_path.parent.mkdir(parents=True, exist_ok=True)
    if capture_path.exists():
        workbook = load_workbook(capture_path)
        worksheet = workbook.active
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "INBOX"
        worksheet.append(CAPTURE_HEADERS)

    try:
        timestamp = datetime.now().astimezone().isoformat(timespec="seconds")
        worksheet.append((timestamp, kind, normalized_title, body))
        workbook.save(capture_path)
    finally:
        workbook.close()
    return f"Excel INBOXへ保存しました: {normalized_title}\n{capture_path}"


def escape_ics_text(text: str) -> str:
    """ICSのテキスト値をRFC 5545に従ってエスケープする。

    Args:
        text: SUMMARYやDESCRIPTIONへ埋め込むテキスト。
    """
    return (
        text.replace("\\", "\\\\")
        .replace(";", "\\;")
        .replace(",", "\\,")
        .replace("\r\n", "\\n")
        .replace("\r", "\\n")
        .replace("\n", "\\n")
    )


def format_calendar_ics(
    subject: str,
    start_time: str,
    body: str,
    uid: str,
    dtstamp: str,
) -> str:
    """予定1件分のICS本文を組み立てる（副作用なし）。

    Args:
        subject: 予定の件名。
        start_time: 開始日時（ISO 8601形式、例: 2026-07-01T14:00:00）。
        body: 予定の本文。空文字の場合はDESCRIPTIONを付けない。
        uid: 予定を一意に識別するUID。
        dtstamp: 生成時刻（UTC、YYYYMMDDTHHMMSSZ形式）。
    """
    start_dt = datetime.fromisoformat(start_time)
    end_dt = start_dt + timedelta(minutes=15)

    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//dev-slm-shortcut//JP",
        "BEGIN:VEVENT",
        f"UID:{uid}",
        f"DTSTAMP:{dtstamp}",
        f"DTSTART:{start_dt.strftime('%Y%m%dT%H%M%S')}",
        f"DTEND:{end_dt.strftime('%Y%m%dT%H%M%S')}",
        f"SUMMARY:{escape_ics_text(subject)}",
    ]
    if body:
        lines.append(f"DESCRIPTION:{escape_ics_text(body)}")
    lines += [
        "BEGIN:VALARM",
        "TRIGGER:PT0S",
        "ACTION:DISPLAY",
        "DESCRIPTION:Reminder",
        "END:VALARM",
        "END:VEVENT",
        "END:VCALENDAR",
        "",
    ]
    return "\r\n".join(lines)


def try_delete_ics(path: Path, attempts: int = 3, delay: float = 2.0) -> bool:
    """カレンダーアプリが読み込む間、数回リトライして一時ICSを削除する。

    Args:
        path: 削除するICSファイル。
        attempts: 試行回数。
        delay: 試行間隔（秒）。
    """
    for _ in range(attempts):
        try:
            path.unlink()
            return True
        except OSError:
            time.sleep(delay)
    return False


def register_pending_cleanup(path: Path) -> None:
    """削除できなかったICSを次回起動時の掃除対象として記録する。

    Args:
        path: 後で削除するICSファイル。
    """
    with CLEANUP_LIST_PATH.open("a", encoding="utf-8") as cleanup_list:
        cleanup_list.write(str(path) + "\n")


def run_pending_cleanup() -> None:
    """前回削除できなかったICS一時ファイルをまとめて削除する。"""
    if not CLEANUP_LIST_PATH.exists():
        return

    remaining = []
    for line in CLEANUP_LIST_PATH.read_text(encoding="utf-8").splitlines():
        path = Path(line.strip())
        if not path.exists():
            continue
        try:
            path.unlink()
        except OSError:
            remaining.append(line)

    if remaining:
        CLEANUP_LIST_PATH.write_text(
            "\n".join(remaining) + "\n", encoding="utf-8"
        )
    else:
        CLEANUP_LIST_PATH.unlink(missing_ok=True)


def create_calendar_task(
    start_time: str,
    entity_type: CalendarEntityType,
    title: str,
    body: str = "",
) -> str:
    """カレンダーの予定（タスク/リマインダ）をICSで作り、確認画面を開く。

    Args:
        start_time: 開始日時（ISO 8601形式、例: 2026-07-01T14:00:00）。
        entity_type: 予定の種別。タスクまたはリマインダのいずれか。
        title: 予定のタイトル。
        body: 予定の本文。依頼文で指定された本文、またはワークフローで
            直前の操作結果が渡る。両方ある場合は結合済みの文字列が渡る。
    """
    if sys.platform != "win32":
        raise OSError("カレンダー登録はWindowsでのみ利用できます。")

    normalized_title = _normalize_title(title, "タスク名")
    if entity_type not in CALENDAR_ENTITY_TYPES:
        allowed = ", ".join(CALENDAR_ENTITY_TYPES)
        raise ValueError(f"未対応の種別です: {entity_type}（許可値: {allowed}）")

    subject = f"【{entity_type}】{normalized_title}"
    uid = f"{uuid.uuid4()}@dev-slm-shortcut"
    dtstamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    ics_content = format_calendar_ics(subject, start_time, body, uid, dtstamp)

    with tempfile.NamedTemporaryFile(
        suffix=".ics", delete=False, mode="wb"
    ) as ics_file:
        ics_file.write(ics_content.encode("utf-8"))
        temp_path = Path(ics_file.name)

    os.startfile(temp_path)
    time.sleep(5)

    if not try_delete_ics(temp_path):
        register_pending_cleanup(temp_path)

    return f"カレンダーの確認画面を開きました: {subject}"


def get_current_page() -> str:
    """直近のChromeまたはEdgeウィンドウからタイトルとURLを取得する。"""
    if sys.platform != "win32":
        raise OSError("現在ページの取得はWindowsでのみ利用できます。")

    completed = subprocess.run(
        [
            "powershell.exe",
            "-NoProfile",
            "-NonInteractive",
            "-Command",
            CURRENT_PAGE_SCRIPT,
        ],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        check=False,
    )
    if completed.returncode != 0:
        detail = completed.stderr.strip()
        raise OSError(
            "ChromeまたはEdgeのページ情報を取得できませんでした。"
            "ブラウザを開いてから再実行してください。"
            + (f"\n{detail}" if detail else "")
        )

    try:
        payload = json.loads(completed.stdout.strip())
        title = payload["title"].strip()
        url = payload["url"].strip()
    except (
        json.JSONDecodeError,
        KeyError,
        TypeError,
        AttributeError,
    ) as error:
        raise OSError("ブラウザから不正なページ情報が返されました。") from error

    if not title or not url:
        raise OSError("ブラウザから空のページ情報が返されました。")
    return f"{title}\n{url}"


def open_recycle_bin() -> str:
    """Windowsのゴミ箱を開いて内容を表示する。"""
    os.startfile("shell:RecycleBinFolder")
    return "ゴミ箱を開きました。"


def format_bytes(size: int) -> str:
    """バイト数を小数1桁のGiB表記へ変換する。

    Args:
        size: 変換するバイト数。
    """
    return f"{size / (1024 ** 3):.1f} GiB"


def show_system_info() -> str:
    """OSとシステムドライブの容量を取得して表示用テキストを返す。"""
    if sys.platform != "win32":
        raise OSError("システム情報の表示はWindowsでのみ利用できます。")

    release, version, _, _ = platform.win32_ver()
    edition = platform.win32_edition()
    build = sys.getwindowsversion().build
    system_drive = os.environ.get("SystemDrive", "C:")
    usage = shutil.disk_usage(system_drive + "\\")

    os_name = " ".join(part for part in ("Windows", edition, release) if part)
    return "\n".join(
        [
            f"OS: {os_name}",
            f"バージョン: {version or '不明'}（ビルド {build}）",
            f"コンピューター名: {platform.node() or '不明'}",
            f"システムドライブ: {system_drive}",
            f"総容量: {format_bytes(usage.total)}",
            f"使用量: {format_bytes(usage.used)}",
            f"空き容量: {format_bytes(usage.free)}",
        ]
    )


def lock_pc() -> str:
    """確認を挟まず現在のWindowsセッションをロックする。"""
    if sys.platform != "win32":
        raise OSError("PCのロックはWindowsでのみ利用できます。")

    if not ctypes.windll.user32.LockWorkStation():
        raise OSError("PCをロックできませんでした。")
    return "PCをロックしました。"


def empty_recycle_bin() -> str:
    """Windowsのゴミ箱にある項目を完全に削除する。"""
    if sys.platform != "win32":
        raise OSError("ゴミ箱を空にする操作はWindowsでのみ利用できます。")

    result = subprocess.run(
        [
            "powershell",
            "-NoProfile",
            "-Command",
            "Clear-RecycleBin -Force -ErrorAction SilentlyContinue",
        ],
        capture_output=True,
        text=True,
        encoding="utf-8",
    )
    if result.stderr.strip():
        raise OSError(f"ゴミ箱を空にできませんでした: {result.stderr.strip()}")
    return "ゴミ箱を空にしました。"


def clear_temp_files() -> str:
    """%TMP%直下の削除可能なファイルとフォルダをすべて削除する。"""
    if sys.platform != "win32":
        raise OSError("一時ファイルの削除はWindowsでのみ利用できます。")

    temp_directory = Path(tempfile.gettempdir())
    if not temp_directory.is_dir():
        raise OSError(f"一時フォルダが見つかりません: {temp_directory}")

    removed = 0
    skipped = 0
    for entry in temp_directory.iterdir():
        try:
            if entry.is_dir() and not entry.is_symlink():
                shutil.rmtree(entry, ignore_errors=True)
            else:
                entry.unlink(missing_ok=True)
        except OSError:
            pass

        if entry.exists():
            skipped += 1
        else:
            removed += 1

    return (
        f"一時フォルダを掃除しました（パス: {temp_directory}、"
        f"削除: {removed}件、スキップ: {skipped}件）。"
    )


def copilot_chat(message: str, model: str = "Opus") -> str:
    """M365 Copilotのチャット欄にメッセージを送信する。

    Args:
        message: 送信するメッセージ。日本語を保持したまま渡す。
        model: 使用するモデル。Opus、Think Deeper、Quick Response、自動など。
    """
    if sys.platform != "win32":
        raise OSError("Copilotチャットの送信はWindowsでのみ利用できます。")
    if not message.strip():
        raise ValueError("送信するメッセージが空です。")

    os.startfile(COPILOT_URL)

    environment = os.environ.copy()
    environment["LOCAL_ACTIONS_COPILOT_MESSAGE"] = message
    environment["LOCAL_ACTIONS_COPILOT_MODEL"] = model

    completed = subprocess.run(
        [
            "powershell.exe",
            "-NoProfile",
            "-NonInteractive",
            "-Sta",
            "-Command",
            COPILOT_CHAT_SCRIPT,
        ],
        env=environment,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        check=False,
    )
    if completed.returncode != 0:
        detail = completed.stderr.strip()
        raise OSError(
            "Copilotへのメッセージ送信に失敗しました。"
            + (f"\n{detail}" if detail else "")
        )
    return completed.stdout.strip() or f"Copilot（モデル: {model}）にメッセージを送信しました。"


def prune_docker_and_compact_wsl() -> str:
    """Dockerの不要データを削除し、固定パスのWSL仮想ディスクを圧縮する。"""
    if sys.platform != "win32":
        raise OSError(
            "Dockerの掃除とWSL仮想ディスクの圧縮は"
            "Windowsでのみ利用できます。"
        )
    if not WSL_VIRTUAL_DISK_PATH.is_file():
        raise OSError(
            "WSL仮想ディスクが見つかりません: "
            f"{WSL_VIRTUAL_DISK_PATH}"
        )

    docker_commands = (
        [
            "wsl.exe",
            "--exec",
            "docker",
            "image",
            "prune",
            "--all",
            "--force",
        ],
        [
            "wsl.exe",
            "--exec",
            "docker",
            "volume",
            "prune",
            "--force",
        ],
        [
            "wsl.exe",
            "--exec",
            "docker",
            "builder",
            "prune",
            "--force",
        ],
    )
    for command in docker_commands:
        completed = subprocess.run(
            command,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
            check=False,
        )
        if completed.returncode != 0:
            detail = completed.stderr.strip()
            raise OSError(
                f"WSL内のDockerの掃除に失敗しました: {' '.join(command)}"
                + (f"\n{detail}" if detail else "")
            )

    shutdown = subprocess.run(
        ["wsl.exe", "--shutdown"],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        check=False,
    )
    if shutdown.returncode != 0:
        detail = shutdown.stderr.strip()
        raise OSError(
            "WSLを停止できませんでした。"
            + (f"\n{detail}" if detail else "")
        )

    diskpart_commands = "\n".join(
        [
            f'select vdisk file="{WSL_VIRTUAL_DISK_PATH}"',
            "attach vdisk readonly",
            "compact vdisk",
            "detach vdisk",
            "exit",
            "",
        ]
    )
    script_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            suffix=".txt",
            delete=False,
            mode="w",
            encoding="utf-8",
            newline="\r\n",
        ) as script:
            script.write(diskpart_commands)
            script_path = Path(script.name)

        environment = os.environ.copy()
        environment["LOCAL_ACTIONS_DISKPART_SCRIPT"] = str(script_path)
        completed = subprocess.run(
            [
                "powershell.exe",
                "-NoProfile",
                "-NonInteractive",
                "-Command",
                ELEVATED_DISKPART_SCRIPT,
            ],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            env=environment,
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
            check=False,
        )
        if completed.returncode != 0:
            detail = completed.stderr.strip()
            raise OSError(
                "WSL仮想ディスクを圧縮できませんでした。"
                "管理者権限の確認を許可してから再実行してください。"
                + (f"\n{detail}" if detail else "")
            )
    finally:
        if script_path is not None:
            script_path.unlink(missing_ok=True)

    return (
        "Dockerの不要データを削除し、WSL仮想ディスクを圧縮しました。"
        f"\n{WSL_VIRTUAL_DISK_PATH}"
    )
